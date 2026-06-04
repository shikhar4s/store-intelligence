from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

from app.layout_store import DEFAULT_LAYOUT, canonical_store_id


def discover_layout_file(root: Path = Path(".")) -> Optional[Path]:
    candidates = []
    for pattern in ("**/store_layout.json", "**/*layout*.json", "**/*layout*.xlsx", "**/*Store layout*.xlsx"):
        candidates.extend(root.glob(pattern))
    filtered = [path for path in candidates if ".venv" not in path.parts and path.is_file()]
    return sorted(filtered, key=lambda p: (0 if p.suffix.lower() == ".json" else 1, len(str(p))))[0] if filtered else None


def load_layout(path: Optional[Path] = None) -> Dict[str, Any]:
    source = path or discover_layout_file()
    layout = dict(DEFAULT_LAYOUT)
    layout["resource_discovery"] = {"layout_file": str(source) if source else None}
    if source is None:
        layout["assumptions"] = ["No layout file found; using normalized default zones."]
        return layout
    if source.suffix.lower() == ".json":
        try:
            loaded = json.loads(source.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                loaded.setdefault("resource_discovery", {})["layout_file"] = str(source)
                return loaded
        except (OSError, json.JSONDecodeError):
            pass
    if source.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(source, data_only=True)
            ws = wb.active
            labels = [
                str(cell.value)
                for row in ws.iter_rows()
                for cell in row
                if cell.value is not None and str(cell.value).strip()
            ]
            layout["resource_discovery"] = {
                "layout_file": str(source),
                "excel_sheets": wb.sheetnames,
                "embedded_images": len(getattr(ws, "_images", [])),
                "cell_labels": labels[:20],
            }
            layout["assumptions"] = [
                "Layout workbook contains visual floor-plan data rather than machine-readable polygons.",
                "Generated normalized zones should be manually adjusted in configs/cameras.example.yaml for production use.",
            ]
        except Exception as exc:
            layout["assumptions"] = [f"Could not parse Excel layout ({exc}); using normalized defaults."]
    return layout


def write_generated_layout(path: Path, layout: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(layout, indent=2, sort_keys=True), encoding="utf-8")
